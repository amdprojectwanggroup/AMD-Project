'''
1. Used to generate decision tree
2. Used for creating adjaency matrix of the fingerprint vectors for molecules.
'''
from keras.models import load_model, Model
import numpy as np
from openpyxl import load_workbook
import os
# Own Scripts
from own_package.features_labels_setup import read_reaction_data, read_reaction_data_smiles
from own_package.SVM_DT_setup import  DT_classifer, create_dt_hparams
from own_package.others import print_array_to_excel
# NGF scripts
from own_package.NGF.layers import NeuralGraphOutput, NeuralGraphHidden


def generate_DT(save_name, loader_file, dt_hparams):
    fl = read_reaction_data(loader_file, mode='c', save_mode=False)
    # in the loader file, put the first 5 examples as the original dataset first 5 examples to see if it learns the
    # temperature variation
    eval_fl, train_fl = fl.create_subsetsplit([0, 5])
    DT = DT_classifer(dt_hparams)
    DT.train_model(train_fl, save_mode=True, plot_mode=True, save_name=save_name)
    predictions_class, acc, cm, f1s, mcc = DT.eval(eval_fl)
    for item in predictions_class, acc, cm, f1s, mcc:
        print(item)


def run_generate_DT():
    dt_hparams = create_dt_hparams(max_depth=4, min_samples_split=300)
    generate_DT('DT_Au25', './excel/data_loader/gold_data_loader_DT_Au25', dt_hparams)




def fp_distance(model_store_dir, loader_excel_file, fp_excel_file):
    fp_store = []
    model_store = []
    for idx, file in enumerate(os.listdir(model_store_dir)):
        filename = os.fsdecode(file)
        if filename.endswith(".h5"):
            model_store.append(model_store_dir + '/' + filename)
    print('Loading models from {}. Total models = {}'.format(model_store_dir, len(model_store)))

    fl = read_reaction_data_smiles(loader_excel_file=loader_excel_file)
    # fl.smiles is an ndarray shape (no. examples, no. molecules)
    smiles_store = []
    idx_store = []
    for idx, single_molecule in enumerate(fl.smiles.T):
        # single_molecule shape (no. of examples,)
        single_dic = {}
        for row_idx, single_example in enumerate(single_molecule):
            if single_example not in single_dic:
                # row_idx is the row number the single_example is in the data_loader.
                # when comparing row_idx to the excel row number, excel row = row_idx + 2
                # since excel row counting starts from 1, and excel 1st row is header, so total is plus 2
                # this is to check if the single_example smiles is not in the dic, add it to the dic along with the idx
                single_dic[single_example] = row_idx
        keys, values = zip(*single_dic.items())
        smiles_store.append(list(keys))
        idx_store.append(list(values))

    input_abe_tensors = []
    for molecule_idx, single_molecule in enumerate(fl.features_d_a):
        tensors = []
        for single_tensor in single_molecule:
            # single_tensor is ndarray shape (no. of total examples, ...)
            # tensors is ndarray shape (no. of unique smiles examples, ...)
            # tensors contains  a, b, e tensor information, where each a,b,e first dim is example idx.
            # input_abe_tensor nested list shape (no. of molecules, 3 abe tensor each is a ndarray)
            tensors.append(single_tensor[idx_store[molecule_idx], ...])
        input_abe_tensors.append(tensors)

    fp_store = [[] for _ in range(fl.features_d_count)]
    for model in model_store:
        model_1 = load_model(model, custom_objects={'NeuralGraphHidden': NeuralGraphHidden,
                                                    'NeuralGraphOutput': NeuralGraphOutput})
        single_model_fp_store = []
        for idx in range(fl.features_d_count):
            # idx represents the molecular idx count. Looping through each molecule abe tensor.
            fp_model = Model(inputs=model_1.get_layer('h_fp_' + str(idx)).get_input_at(0),
                             outputs=model_1.get_layer('h_fp_' + str(idx)).get_output_at(0))
            # single_model_fp_store list of shape (no. of molecules, ndarray of shape (no. of unique examples, ...) )
            single_model_fp_store.append(fp_model.predict(input_abe_tensors[idx]))

        try:
            for idx, _ in enumerate(fp_store):
                # sum up ndarray for each model instance
                # fp_store has same shape as single_model_fp_store, except it is the sum of all model instances
                fp_store[idx] += single_model_fp_store[idx]
        except ValueError:
            # Means fp_store is still empty. Store ndarray into fp_store for the first iteration.
            for idx, _ in enumerate(fp_store):
                fp_store[idx] = single_model_fp_store[idx]

    # Preparing to write to excel
    wb = load_workbook(fp_excel_file)
    wb.create_sheet('fp')
    sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]
    excel_row = 1
    excel_col = 1

    # Calculating fp dist adjacency matrix and then writing to excel
    for molecule_idx, single_molecule in enumerate(fp_store):
        single_molecule = np.divide(single_molecule, len(model_store))

        x, y = single_molecule.shape
        fp_dist = np.zeros(shape=(x, x))
        for i in range(x):
            cols = range(i + 1, x)
            for j in cols:
                dist = np.linalg.norm(single_molecule[i, :] - single_molecule[j, :])
                fp_dist[i, j] = dist

        ws.cell(excel_row, excel_col).value = 'fp_dist_' + str(molecule_idx)
        print_array_to_excel(array=smiles_store[molecule_idx], first_cell=(excel_row + 1, excel_col), ws=ws, axis=0)
        print_array_to_excel(array=smiles_store[molecule_idx], first_cell=(excel_row, excel_col + 1), ws=ws, axis=1)
        print_array_to_excel(array=fp_dist, first_cell=(excel_row + 1, excel_col + 1), ws=ws, axis=2)

        ws.cell(excel_row + x + 2, excel_col).value = 'fp_' + str(molecule_idx) + '_actual_values'
        print_array_to_excel(array=smiles_store[molecule_idx], first_cell=(excel_row + x + 3, excel_col), ws=ws, axis=0)
        print_array_to_excel(array=single_molecule, first_cell=(excel_row + x + 3, excel_col + 1), ws=ws, axis=2)
        excel_col += max(y, single_molecule.shape[0]) + 2

    wb.save(fp_excel_file)
    wb.close()

if __name__ == '__main__':
    fp_distance(model_store_dir='./save/models/final/SNN_smiles_desc_all',
                loader_excel_file='./excel/data_loader/gold_data_loader.xlsm', fp_excel_file='./excel/fp/fp.xlsx')
